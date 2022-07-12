from openpyxl import Workbook

# 엑셀 시트 관리
wb = Workbook()
ws = wb.create_sheet() # 새로운 Sheet를 기본 이름 생성
ws.title = 'MySheet'

ws1 = wb.create_sheet('Your Sheet') # 주어진 이름으로 시트생성
ws2 = wb.create_sheet('NewSheet',2) # 시트 index 2번째에 시트생성

# 시트 접근 ws1,ws2,.. 처럼 하는 방법도 있고, wb['시트명']처럼 dict 형태로도 접근 가능
print(wb['NewSheet'].title)
new_ws = wb['NewSheet']
print(new_ws)

# 모든 시트 확인, 리스트로 반환
print(wb.sheetnames)

# Sheet 복사
new_ws['A1'] = 'Test' #A1 셀에 데이터 넣음
target = wb.copy_worksheet(new_ws) # 복사된 Sheet가 우측 마지막에 생성됨(데이터 포함)
target.title = 'Copied_Sheet'


wb.save('nadoCoding_sample2.xlsx')